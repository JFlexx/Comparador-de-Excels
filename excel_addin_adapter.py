from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from comparator import (
    ComparisonRequest,
    DECISION_COLUMNS,
    DecisionsLoadRequest,
    MergeRequest,
    SERVICE,
    source_action_for_base,
)
from excel_integration_contracts import (
    ExcelCompareContract,
    ExcelComparisonResult,
    ExcelDecisionRow,
    ExcelDecisionSheetContract,
    ExcelDecisionSheetLoadResult,
    ExcelDecisionsReadResult,
    ExcelMergeContract,
    ExcelMergeResult,
    ExcelSheetSummary,
    ExcelWorkbookSelection,
)


class ExcelAddinAdapter:
    """Adaptador Excel-first para Excel Desktop vía xlwings/Ribbon + Python local.

    El add-in solo intercambia contratos serializables (dict/JSON-friendly) con este adaptador;
    `comparator.py` sigue siendo el motor único de comparación y merge.
    """

    integration_route = "Excel Desktop via xlwings Ribbon + local Python runtime"

    def select_workbooks(
        self,
        *,
        base_workbook_path: str | Path,
        source_workbook_path: str | Path,
        base_side: str = "a",
    ) -> ExcelWorkbookSelection:
        return ExcelWorkbookSelection.create(
            base_workbook_path=base_workbook_path,
            source_workbook_path=source_workbook_path,
            base_side=base_side,
        )

    def select_workbooks_from_payload(self, payload: dict[str, Any]) -> dict[str, Any]:
        selection = ExcelWorkbookSelection.from_dict(payload)
        return selection.to_dict()

    def compare(self, contract: ExcelCompareContract) -> ExcelComparisonResult:
        diff = SERVICE.compare(
            ComparisonRequest(
                path_a=self._workbook_path_for_side(contract.selection, "a"),
                path_b=self._workbook_path_for_side(contract.selection, "b"),
                options=contract.to_compare_options(),
            )
        )
        default_action = source_action_for_base(contract.selection.base_side)
        decisions_df = diff.to_dataframe(default_action=default_action)

        summary_rows = [ExcelSheetSummary.from_dict(row) for row in diff.summary_rows()]
        decision_rows = [
            ExcelDecisionRow.from_dict(record)
            for record in decisions_df.to_dict(orient="records")
        ]

        if contract.selection.base_side == "a":
            only_in_base = diff.only_in_a
            only_in_source = diff.only_in_b
        else:
            only_in_base = diff.only_in_b
            only_in_source = diff.only_in_a

        return ExcelComparisonResult(
            route=self.integration_route,
            selection=contract.selection,
            compare_mode=contract.compare_mode,
            header_row=contract.header_row,
            total_differences=diff.total_differences,
            common_sheets=list(diff.common_sheets),
            only_in_base=list(only_in_base),
            only_in_source=list(only_in_source),
            default_action=default_action,
            summary_rows=summary_rows,
            decision_rows=decision_rows,
        )

    def compare_payload(self, payload: dict[str, Any]) -> dict[str, Any]:
        return self.compare(ExcelCompareContract.from_dict(payload)).to_dict()

    def load_decision_table_into_workbook(
        self,
        comparison: ExcelComparisonResult,
        target: ExcelDecisionSheetContract,
    ) -> ExcelDecisionSheetLoadResult:
        workbook_path = Path(target.workbook_path)
        wb = load_workbook(workbook_path)

        if target.sheet_name in wb.sheetnames:
            ws = wb[target.sheet_name]
            if target.clear_sheet:
                ws.delete_rows(1, ws.max_row)
        else:
            ws = wb.create_sheet(target.sheet_name)

        ws.append(DECISION_COLUMNS)
        for decision in comparison.decision_rows:
            row = decision.to_dict()
            ws.append([row[column] for column in DECISION_COLUMNS])

        wb.save(workbook_path)
        return ExcelDecisionSheetLoadResult(
            workbook_path=str(workbook_path),
            sheet_name=target.sheet_name,
            rows_written=len(comparison.decision_rows),
        )

    def load_decision_table_payload(
        self,
        comparison_payload: dict[str, Any],
        target_payload: dict[str, Any],
    ) -> dict[str, Any]:
        result = self.load_decision_table_into_workbook(
            ExcelComparisonResult.from_dict(comparison_payload),
            ExcelDecisionSheetContract.from_dict(target_payload),
        )
        return result.to_dict()

    def read_decisions_from_workbook(self, target: ExcelDecisionSheetContract) -> ExcelDecisionsReadResult:
        decisions_df = SERVICE.decisions_from_excel(
            DecisionsLoadRequest(path=target.workbook_path, sheet_name=target.sheet_name)
        )
        decision_rows = [
            ExcelDecisionRow.from_dict(record)
            for record in decisions_df.to_dict(orient="records")
        ]
        return ExcelDecisionsReadResult(
            sheet_name=target.sheet_name,
            rows_loaded=len(decision_rows),
            decisions=decision_rows,
        )

    def read_decisions_payload(self, target_payload: dict[str, Any]) -> dict[str, Any]:
        return self.read_decisions_from_workbook(
            ExcelDecisionSheetContract.from_dict(target_payload)
        ).to_dict()

    def execute_merge(self, contract: ExcelMergeContract) -> ExcelMergeResult:
        decisions_df = SERVICE.decisions_from_excel(
            DecisionsLoadRequest(
                path=contract.decisions_workbook_path,
                sheet_name=contract.decisions_sheet_name,
            )
        )
        output = SERVICE.apply_decisions(
            MergeRequest(
                workbook_a=self._workbook_path_for_side(contract.selection, "a"),
                workbook_b=self._workbook_path_for_side(contract.selection, "b"),
                decisions=decisions_df,
                output_path=contract.output_path,
                base=contract.selection.base_side,
                include_sheets_from_source_only=contract.include_sheets_from_source_only,
            )
        )
        return ExcelMergeResult(
            output_path=str(output),
            base_workbook_path=contract.selection.base_workbook_path,
            source_workbook_path=contract.selection.source_workbook_path,
            decisions_sheet_name=contract.decisions_sheet_name,
        )

    def execute_merge_payload(self, payload: dict[str, Any]) -> dict[str, Any]:
        return self.execute_merge(ExcelMergeContract.from_dict(payload)).to_dict()

    @staticmethod
    def _workbook_path_for_side(selection: ExcelWorkbookSelection, side: str) -> str:
        if selection.base_side == "a":
            return selection.base_workbook_path if side == "a" else selection.source_workbook_path
        return selection.source_workbook_path if side == "a" else selection.base_workbook_path
