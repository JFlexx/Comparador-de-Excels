from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
import hashlib
from pathlib import Path
from typing import Dict, Iterable, List, Literal, Mapping, Optional, Sequence
from urllib.parse import quote, unquote

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

WorkbookSide = Literal["a", "b"]
CompareMode = Literal["coordinate", "row-based"]

DEFAULT_ACTION = "use_b"
VALID_ACTIONS = {"use_a", "use_b", "manual"}
VALID_COMPARE_MODES = {"coordinate", "row-based"}
DECISION_SHEET_NAME = "Decisiones"
DECISION_METADATA_SHEET_NAME = "_metadata"
DECISION_FORMAT_VERSION = "2.0"
DECISION_COLUMNS = [
    "decision_id",
    "sheet",
    "cell",
    "row",
    "column",
    "column_letter",
    "context",
    "header",
    "key",
    "diff_type",
    "value_a",
    "value_b",
    "action",
    "manual_value",
    "reviewed",
]
EDITABLE_DECISION_COLUMNS = {"action", "manual_value", "reviewed"}


@dataclass(frozen=True)
class ComparisonRequest:
    """Contrato de entrada para comparar dos libros Excel.

    path_a/path_b:
        Rutas a archivos `.xlsx` o `.xlsm` accesibles por el proceso.
    options:
        Reglas de comparación. En `coordinate`, se compara posición exacta.
        En `row-based`, se usan encabezados y `sheet_keys` para identificar registros.
    """

    path_a: str | Path
    path_b: str | Path
    options: "CompareOptions" = field(default_factory=lambda: CompareOptions())


@dataclass(frozen=True)
class DecisionTemplateRequest:
    """Contrato para exportar una plantilla Excel editable de decisiones."""

    diff: "WorkbookDiff"
    output_path: str | Path
    default_action: str = DEFAULT_ACTION


@dataclass(frozen=True)
class DecisionsLoadRequest:
    """Contrato para leer una plantilla de decisiones desde Excel."""

    path: str | Path
    sheet_name: str = DECISION_SHEET_NAME


@dataclass(frozen=True)
class MergeRequest:
    """Contrato para aplicar decisiones y producir un libro final.

    workbook_a/workbook_b:
        Rutas a los dos libros originales.
    decisions:
        DataFrame con columnas al menos `decision_id`, `sheet`, `row`, `column`, `action`.
        Para `manual`, también se usa `manual_value`.
    base:
        `"a"` o `"b"`; indica sobre qué libro se construye el resultado final.
    include_sheets_from_source_only:
        Si es `True`, copia hojas que existan solo en el libro origen.
    compare_mode/header_row/sheet_keys:
        Metadata del diff original. En `row-based`, el merge usa `key`, `header`
        y `diff_type` para ubicar registros por identidad lógica, no solo por coordenada.
    """

    workbook_a: str | Path
    workbook_b: str | Path
    decisions: pd.DataFrame
    output_path: str | Path
    base: WorkbookSide = "a"
    include_sheets_from_source_only: bool = True
    compare_mode: CompareMode = "coordinate"
    header_row: int = 1
    sheet_keys: Dict[str, List[str]] = field(default_factory=dict)


class ComparatorService:
    """API estable del motor para futuras interfaces corporativas.

    Las interfaces deben invocar esta clase o las funciones públicas homónimas de este módulo,
    evitando duplicar lógica de negocio.
    """

    def compare(self, request: ComparisonRequest) -> "WorkbookDiff":
        return compare_workbooks(request.path_a, request.path_b, options=request.options)

    def export_decision_template(self, request: DecisionTemplateRequest) -> Path:
        return export_decision_template(
            diff=request.diff,
            output_path=request.output_path,
            default_action=request.default_action,
        )

    def decisions_from_excel(self, request: DecisionsLoadRequest) -> pd.DataFrame:
        return decisions_from_excel(path=request.path, sheet_name=request.sheet_name)

    def apply_decisions(self, request: MergeRequest) -> Path:
        return apply_decisions(
            workbook_a=request.workbook_a,
            workbook_b=request.workbook_b,
            decisions=request.decisions,
            output_path=request.output_path,
            base=request.base,
            include_sheets_from_source_only=request.include_sheets_from_source_only,
            compare_mode=request.compare_mode,
            header_row=request.header_row,
            sheet_keys=request.sheet_keys,
        )


SERVICE = ComparatorService()


@dataclass
class CompareOptions:
    """Opciones de comparación del núcleo.

    Contrato:
    - `compare_mode="coordinate"`: diff por coordenada exacta fila/columna.
    - `compare_mode="row-based"`: diff por registro usando fila de encabezados + `sheet_keys`.
    - `sheet_keys`: `{"Hoja": ["ColumnaId", "OtraClave"]}`.
    - `header_row`: índice 1-based de la fila que contiene encabezados.
    """

    strip_strings: bool = True
    case_sensitive: bool = True
    ignore_empty_string_vs_none: bool = True
    compare_mode: CompareMode = "coordinate"
    sheet_keys: Dict[str, List[str]] = field(default_factory=dict)
    header_row: int = 1

    def __post_init__(self) -> None:
        if self.compare_mode not in VALID_COMPARE_MODES:
            raise ValueError(f"compare_mode debe ser uno de {sorted(VALID_COMPARE_MODES)}")
        self.sheet_keys = {sheet: list(keys) for sheet, keys in self.sheet_keys.items()}
        if self.header_row < 1:
            raise ValueError("header_row debe ser >= 1")


@dataclass
class CellDiff:
    sheet: str
    row: int
    column: int
    value_a: Optional[object]
    value_b: Optional[object]
    diff_type: str = "modified"
    key: Optional[str] = None
    header: Optional[str] = None

    @property
    def coordinate(self) -> str:
        return column_letter(self.column) + str(self.row)


@dataclass
class SheetDiffSummary:
    sheet: str
    differences: List[CellDiff]
    total_differences: int = field(init=False)
    columns: List[str] = field(init=False)
    row_numbers: List[int] = field(init=False)

    def __post_init__(self) -> None:
        self.total_differences = len(self.differences)
        self.columns = sorted({column_letter(diff.column) for diff in self.differences})
        self.row_numbers = sorted({diff.row for diff in self.differences})


@dataclass
class WorkbookDiff:
    """Resultado estable de `compare_workbooks`.

    Salida principal para UI, CLI o futuros add-ins.
    - `only_in_a` / `only_in_b`: hojas exclusivas por libro.
    - `common_sheets`: hojas comparadas.
    - `differences`: mapa `sheet -> list[CellDiff]`.
    - `to_dataframe()`: DataFrame estándar para revisión/merge.
    """

    only_in_a: List[str]
    only_in_b: List[str]
    common_sheets: List[str]
    differences: Dict[str, List[CellDiff]]
    options: CompareOptions
    workbook_a: Optional[str] = None
    workbook_b: Optional[str] = None
    grouped_differences: Dict[str, SheetDiffSummary] = field(init=False)
    total_differences: int = field(init=False)
    _all_differences: List[CellDiff] = field(init=False, repr=False)

    def __post_init__(self) -> None:
        self.grouped_differences = {}
        self._all_differences = []

        for sheet in self.common_sheets:
            sheet_diffs = list(self.differences.get(sheet, []))
            self.differences[sheet] = sheet_diffs
            self.grouped_differences[sheet] = SheetDiffSummary(sheet=sheet, differences=sheet_diffs)
            self._all_differences.extend(sheet_diffs)

        self.total_differences = len(self._all_differences)

    def all_differences(self) -> List[CellDiff]:
        return list(self._all_differences)

    def summary_rows(self) -> List[dict[str, object]]:
        return [
            {
                "sheet": summary.sheet,
                "differences": summary.total_differences,
                "columns": ", ".join(summary.columns),
                "rows": len(summary.row_numbers),
            }
            for summary in self.grouped_differences.values()
        ]

    def to_dataframe(self, default_action: str = DEFAULT_ACTION) -> pd.DataFrame:
        dataframe = diffs_to_dataframe(self._all_differences, default_action=default_action)
        _attach_decisions_metadata(
            dataframe,
            compare_mode=self.options.compare_mode,
            header_row=self.options.header_row,
            sheet_keys=self.options.sheet_keys,
        )
        return dataframe


def column_letter(column: int) -> str:
    letters = ""
    current = column
    while current > 0:
        current, rem = divmod(current - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def _normalize(value: object, options: CompareOptions) -> object:
    if isinstance(value, str):
        normalized = value.strip() if options.strip_strings else value
        return normalized if options.case_sensitive else normalized.lower()

    if options.ignore_empty_string_vs_none and value == "":
        return None

    return value


def _stringify_key_part(value: object) -> str:
    return "<vacío>" if value is None else str(value)


def _build_key_label(key_headers: Sequence[str], key_values: tuple[object, ...], fallback_row: int) -> str:
    if key_headers:
        parts = [f"{header}={_stringify_key_part(value)}" for header, value in zip(key_headers, key_values)]
        return ", ".join(parts)
    return f"fila={fallback_row}"


def _read_sheet_rows(ws: Worksheet, options: CompareOptions) -> dict[str, object]:
    header_row = options.header_row
    raw_headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]

    headers: List[str] = []
    header_to_column: Dict[str, int] = {}
    for idx, raw_header in enumerate(raw_headers, start=1):
        header = str(raw_header).strip() if raw_header is not None else f"__col_{idx}"
        if not header:
            header = f"__col_{idx}"
        if header in header_to_column:
            header = f"{header}__{idx}"
        headers.append(header)
        header_to_column[header] = idx

    rows: List[dict[str, object]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        values_by_header = {
            header: ws.cell(row=row_idx, column=col_idx).value
            for col_idx, header in enumerate(headers, start=1)
        }
        if all(value is None for value in values_by_header.values()):
            continue
        rows.append({"row_idx": row_idx, "values": values_by_header})

    return {"headers": headers, "header_to_column": header_to_column, "rows": rows}


def _record_key(
    sheet_name: str,
    record: Mapping[str, object],
    key_headers: Sequence[str],
    comparable_headers: Sequence[str],
    options: CompareOptions,
) -> tuple[object, ...]:
    values = record["values"]
    if not isinstance(values, Mapping):
        raise ValueError(f"La fila de la hoja '{sheet_name}' no tiene una estructura válida")

    if key_headers:
        missing = [header for header in key_headers if header not in values]
        if missing:
            raise ValueError(f"La hoja '{sheet_name}' no contiene las columnas clave requeridas: {missing}")
        return tuple(_normalize(values[header], options) for header in key_headers)

    return tuple((header, _normalize(values.get(header), options)) for header in comparable_headers)


def _compare_sheet_by_rows(sheet_name: str, ws_a: Worksheet, ws_b: Worksheet, options: CompareOptions) -> List[CellDiff]:
    parsed_a = _read_sheet_rows(ws_a, options)
    parsed_b = _read_sheet_rows(ws_b, options)

    headers_a = list(parsed_a["headers"])
    headers_b = list(parsed_b["headers"])
    comparable_headers = list(dict.fromkeys([*headers_a, *headers_b]))
    key_headers = options.sheet_keys.get(sheet_name, [])

    grouped_a: dict[tuple[object, ...], List[dict[str, object]]] = defaultdict(list)
    grouped_b: dict[tuple[object, ...], List[dict[str, object]]] = defaultdict(list)

    for record in parsed_a["rows"]:
        grouped_a[_record_key(sheet_name, record, key_headers, comparable_headers, options)].append(record)
    for record in parsed_b["rows"]:
        grouped_b[_record_key(sheet_name, record, key_headers, comparable_headers, options)].append(record)

    diffs: List[CellDiff] = []

    for key in sorted(set(grouped_a) | set(grouped_b), key=lambda item: repr(item)):
        records_a = grouped_a.get(key, [])
        records_b = grouped_b.get(key, [])
        shared_count = min(len(records_a), len(records_b))

        for index in range(shared_count):
            record_a = records_a[index]
            record_b = records_b[index]
            key_label = _build_key_label(key_headers, key, int(record_b["row_idx"]))
            for header in comparable_headers:
                value_a = record_a["values"].get(header)
                value_b = record_b["values"].get(header)
                if _normalize(value_a, options) == _normalize(value_b, options):
                    continue
                column = parsed_b["header_to_column"].get(header) or parsed_a["header_to_column"].get(header) or 1
                diffs.append(
                    CellDiff(
                        sheet=sheet_name,
                        row=int(record_b["row_idx"]),
                        column=int(column),
                        value_a=value_a,
                        value_b=value_b,
                        diff_type="modified",
                        key=key_label,
                        header=header,
                    )
                )

        for record_a in records_a[shared_count:]:
            key_label = _build_key_label(key_headers, key, int(record_a["row_idx"]))
            for header in comparable_headers:
                value_a = record_a["values"].get(header)
                if _normalize(value_a, options) is None:
                    continue
                column = parsed_a["header_to_column"].get(header) or 1
                diffs.append(
                    CellDiff(
                        sheet=sheet_name,
                        row=int(record_a["row_idx"]),
                        column=int(column),
                        value_a=value_a,
                        value_b=None,
                        diff_type="deleted",
                        key=key_label,
                        header=header,
                    )
                )

        for record_b in records_b[shared_count:]:
            key_label = _build_key_label(key_headers, key, int(record_b["row_idx"]))
            for header in comparable_headers:
                value_b = record_b["values"].get(header)
                if _normalize(value_b, options) is None:
                    continue
                column = parsed_b["header_to_column"].get(header) or 1
                diffs.append(
                    CellDiff(
                        sheet=sheet_name,
                        row=int(record_b["row_idx"]),
                        column=int(column),
                        value_a=None,
                        value_b=value_b,
                        diff_type="added",
                        key=key_label,
                        header=header,
                    )
                )

    return diffs


def build_decision_id(diff: CellDiff) -> str:
    payload = "|".join(
        [
            diff.sheet,
            str(diff.row),
            str(diff.column),
            diff.diff_type,
            "" if diff.header is None else str(diff.header),
            "" if diff.key is None else str(diff.key),
            "" if diff.value_a is None else repr(diff.value_a),
            "" if diff.value_b is None else repr(diff.value_b),
        ]
    )
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:12]
    return f"{quote(diff.sheet, safe='')}|{diff.row}|{diff.column}|{digest}"


def _parse_decision_id(decision_id: object) -> tuple[str, int, int]:
    value = "" if decision_id is None else str(decision_id).strip()
    parts = value.split("|")
    if len(parts) != 4:
        raise ValueError("decision_id no tiene el formato esperado")

    try:
        sheet_name = unquote(parts[0])
        row_index = int(parts[1])
        column_index = int(parts[2])
    except (TypeError, ValueError) as exc:
        raise ValueError("decision_id contiene coordenadas inválidas") from exc

    if not sheet_name or row_index < 1 or column_index < 1:
        raise ValueError("decision_id contiene coordenadas inválidas")

    return sheet_name, row_index, column_index


def _target_base_for_default_action(default_action: str) -> str:
    return "b" if default_action == "use_a" else "a"


def _source_signature(diff: WorkbookDiff) -> str:
    parts = [diff.options.compare_mode, str(diff.options.header_row)]
    for workbook_path in (diff.workbook_a, diff.workbook_b):
        if workbook_path:
            path = Path(workbook_path)
            stat = path.stat()
            parts.extend([path.name, str(stat.st_size), str(stat.st_mtime_ns)])
        else:
            parts.append("<unknown>")
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()[:16]


def compare_workbooks(
    path_a: str | Path,
    path_b: str | Path,
    options: CompareOptions | None = None,
) -> WorkbookDiff:
    """Compara dos libros y devuelve un `WorkbookDiff`.

    Entradas:
    - `path_a`, `path_b`: rutas a libros Excel.
    - `options`: `CompareOptions`; si se omite se usa modo `coordinate`.

    Salida:
    - `WorkbookDiff` con hojas exclusivas, hojas comunes y diferencias detalladas.
    """

    options = options or CompareOptions()
    wb_a = load_workbook(path_a, data_only=False)
    wb_b = load_workbook(path_b, data_only=False)

    sheets_a = set(wb_a.sheetnames)
    sheets_b = set(wb_b.sheetnames)

    only_in_a = sorted(sheets_a - sheets_b)
    only_in_b = sorted(sheets_b - sheets_a)
    common = sorted(sheets_a & sheets_b)

    differences: Dict[str, List[CellDiff]] = {}

    for sheet_name in common:
        ws_a = wb_a[sheet_name]
        ws_b = wb_b[sheet_name]

        if options.compare_mode == "row-based":
            differences[sheet_name] = _compare_sheet_by_rows(sheet_name, ws_a, ws_b, options)
            continue

        max_row = max(ws_a.max_row, ws_b.max_row)
        max_col = max(ws_a.max_column, ws_b.max_column)
        sheet_diffs: List[CellDiff] = []

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                val_a = ws_a.cell(row=row, column=col).value
                val_b = ws_b.cell(row=row, column=col).value
                if _normalize(val_a, options) != _normalize(val_b, options):
                    sheet_diffs.append(
                        CellDiff(
                            sheet=sheet_name,
                            row=row,
                            column=col,
                            value_a=val_a,
                            value_b=val_b,
                            diff_type="modified",
                        )
                    )

        differences[sheet_name] = sheet_diffs

    return WorkbookDiff(
        only_in_a=only_in_a,
        only_in_b=only_in_b,
        common_sheets=common,
        differences=differences,
        options=options,
        workbook_a=str(path_a),
        workbook_b=str(path_b),
    )


def diffs_to_dataframe(diffs: Iterable[CellDiff], default_action: str = DEFAULT_ACTION) -> pd.DataFrame:
    """Convierte diferencias a un DataFrame estándar para UI, Excel y APIs internas.

    Contrato de salida:
    - Siempre usa columnas `DECISION_COLUMNS`.
    - `action` acepta `use_a`, `use_b`, `manual`.
    - `row` es 1-based y `column` es índice numérico 1-based.
    - En `row-based`, `header`, `key` y `diff_type` permiten reconstruir el contexto del registro.
    """

    if default_action not in VALID_ACTIONS:
        raise ValueError(f"default_action debe ser uno de {sorted(VALID_ACTIONS)}")

    rows: list[dict[str, object]] = []
    for diff in diffs:
        rows.append(
            {
                "decision_id": build_decision_id(diff),
                "sheet": diff.sheet,
                "cell": diff.coordinate,
                "row": diff.row,
                "column": diff.column,
                "column_letter": column_letter(diff.column),
                "context": f"{diff.sheet}!{diff.coordinate} · fila {diff.row} · columna {column_letter(diff.column)} ({diff.column})",
                "header": diff.header,
                "key": diff.key,
                "diff_type": diff.diff_type,
                "value_a": diff.value_a,
                "value_b": diff.value_b,
                "action": default_action,
                "manual_value": None,
                "reviewed": False,
            }
        )

    return pd.DataFrame(rows, columns=DECISION_COLUMNS)


def _write_metadata_sheet(wb: Workbook, diff: WorkbookDiff, default_action: str) -> None:
    metadata = wb.create_sheet(DECISION_METADATA_SHEET_NAME)
    metadata.sheet_state = "hidden"
    metadata.append(["key", "value"])
    metadata_rows = [
        ("format_version", DECISION_FORMAT_VERSION),
        ("compare_mode", diff.options.compare_mode),
        ("base_workbook", _target_base_for_default_action(default_action)),
        ("generated_at", datetime.now(timezone.utc).isoformat()),
        ("source_signature", _source_signature(diff)),
        ("workbook_a", Path(diff.workbook_a).name if diff.workbook_a else None),
        ("workbook_b", Path(diff.workbook_b).name if diff.workbook_b else None),
    ]
    for key, value in metadata_rows:
        metadata.append([key, value])



def _add_decision_table(ws: Worksheet) -> None:
    if ws.max_row < 2:
        return

    last_column = column_letter(ws.max_column)
    table = Table(displayName="DecisionTable", ref=f"A1:{last_column}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)



def _protect_decision_sheet(ws: Worksheet) -> None:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            header = str(ws.cell(row=1, column=cell.column).value)
            cell.protection = Protection(locked=header not in EDITABLE_DECISION_COLUMNS)

    ws.protection.sheet = True
    ws.protection.autoFilter = True
    ws.protection.sort = True
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True



def _style_decision_sheet(ws: Worksheet) -> None:
    for cell in ws[1]:
        cell.fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)

    dv = DataValidation(type="list", formula1='"use_a,use_b,manual"', allow_blank=False)
    ws.add_data_validation(dv)
    if ws.max_row >= 2:
        action_column = DECISION_COLUMNS.index("action") + 1
        dv.add(f"{column_letter(action_column)}2:{column_letter(action_column)}{ws.max_row}")
    ws.freeze_panes = "A2"
    _add_decision_table(ws)
    _protect_decision_sheet(ws)



def export_decision_template(
    diff: WorkbookDiff,
    output_path: str | Path,
    default_action: str = DEFAULT_ACTION,
) -> Path:
    """Exporta la plantilla estándar `Decisiones` + `Resumen`.

    Entrada:
    - `diff`: `WorkbookDiff` proveniente del motor.
    - `output_path`: ruta destino.
    - `default_action`: una de `VALID_ACTIONS`.

    Salida:
    - `Path` del archivo generado.
    """

    if default_action not in VALID_ACTIONS:
        raise ValueError(f"default_action debe ser uno de {sorted(VALID_ACTIONS)}")

    wb = Workbook()
    ws = wb.active
    ws.title = DECISION_SHEET_NAME
    ws.append(DECISION_COLUMNS)

    for row in diff.to_dataframe(default_action=default_action).itertuples(index=False):
        ws.append(list(row))

    _style_decision_sheet(ws)
    _write_metadata_sheet(wb, diff, default_action)

    summary = wb.create_sheet(SUMMARY_SHEET_NAME)
    summary.append(["Métrica", "Valor"])
    summary.append(["Modo de comparación", diff.options.compare_mode])
    summary.append(["Fila de encabezados", diff.options.header_row])
    summary.append(["Hojas en común", len(diff.common_sheets)])
    summary.append(["Hojas solo en A", len(diff.only_in_a)])
    summary.append(["Hojas solo en B", len(diff.only_in_b)])
    summary.append(["Diferencias", diff.total_differences])
    if diff.only_in_a:
        summary.append(["Lista solo en A", ", ".join(diff.only_in_a)])
    if diff.only_in_b:
        summary.append(["Lista solo en B", ", ".join(diff.only_in_b)])
    for sheet_name, columns in sorted(diff.options.sheet_keys.items()):
        summary.append([f"{SUMMARY_KEYS_PREFIX}{sheet_name}", ", ".join(columns)])
    for row in diff.summary_rows():
        summary.append([f"Diferencias en {row['sheet']}", row["differences"]])

    output = Path(output_path)
    wb.save(output)
    return output


def _empty_decisions_dataframe() -> pd.DataFrame:
    return pd.DataFrame(columns=DECISION_COLUMNS)



def _coerce_reviewed(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "sí", "si", "yes", "x"}



def decisions_from_excel(path: str | Path, sheet_name: str = DECISION_SHEET_NAME) -> pd.DataFrame:
    """Lee y normaliza una hoja de decisiones Excel a DataFrame estándar.

    Salida:
    - DataFrame con `DECISION_COLUMNS`.
    - `action` validada contra `VALID_ACTIONS`.
    - `reviewed` normalizada a boolean.
    """

    wb = load_workbook(path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{sheet_name}' en el archivo de decisiones")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    if not rows:
        return _empty_decisions_dataframe()

    headers = [str(header) if header is not None else "" for header in rows[0]]
    data = rows[1:]
    df = pd.DataFrame(data, columns=headers)
    return validate_decisions_dataframe(df)



def _read_summary_metadata(ws: Worksheet) -> tuple[CompareMode, int, dict[str, list[str]]]:
    compare_mode: CompareMode = "coordinate"
    header_row = 1
    sheet_keys: dict[str, list[str]] = {}

    for metric, value in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if metric == "Modo de comparación" and value in VALID_COMPARE_MODES:
            compare_mode = str(value)
        elif metric == "Fila de encabezados" and value is not None:
            header_row = int(value)
        elif isinstance(metric, str) and metric.startswith(SUMMARY_KEYS_PREFIX) and value:
            sheet_name = metric.removeprefix(SUMMARY_KEYS_PREFIX).strip()
            columns = [column.strip() for column in str(value).split(",") if column.strip()]
            if sheet_name and columns:
                sheet_keys[sheet_name] = columns

    return compare_mode, header_row, sheet_keys


def _resolve_direction(base: WorkbookSide) -> tuple[WorkbookSide, WorkbookSide]:
    if base not in {"a", "b"}:
        raise ValueError("base debe ser 'a' o 'b'")
    return ("a", "b") if base == "a" else ("b", "a")



def source_action_for_base(base: WorkbookSide) -> str:
    """Devuelve la acción por defecto para traer cambios desde el libro fuente al base."""

    _, source_key = _resolve_direction(base)
    return f"use_{source_key}"



def validate_decisions_dataframe(decisions: pd.DataFrame) -> pd.DataFrame:
    """Valida y normaliza el DataFrame de decisiones esperado por el motor."""

    missing = [column for column in DECISION_COLUMNS if column not in decisions.columns]
    if missing:
        raise ValueError(f"El DataFrame de decisiones no contiene columnas requeridas: {missing}")

    normalized = decisions[DECISION_COLUMNS].copy()
    normalized["decision_id"] = normalized["decision_id"].fillna("").astype(str).str.strip()
    normalized["action"] = normalized["action"].fillna(DEFAULT_ACTION).astype(str).str.strip()
    normalized["reviewed"] = normalized["reviewed"].map(_coerce_reviewed).fillna(False)

    for column in ("sheet", "cell", "column_letter", "context", "header", "key", "diff_type"):
        normalized[column] = normalized[column].where(normalized[column].notna(), None)

    invalid_actions = sorted(set(normalized.loc[~normalized["action"].isin(VALID_ACTIONS), "action"].tolist()))
    if invalid_actions:
        raise ValueError(f"Acciones no válidas: {invalid_actions}. Válidas: {sorted(VALID_ACTIONS)}")

    duplicate_ids = sorted(
        set(normalized.loc[normalized["decision_id"].ne("") & normalized["decision_id"].duplicated(), "decision_id"].tolist())
    )
    if duplicate_ids:
        raise ValueError(f"decision_id duplicados: {duplicate_ids}")

    raw_row_values = normalized["row"].copy()
    raw_column_values = normalized["column"].copy()
    normalized["row"] = pd.to_numeric(normalized["row"], errors="coerce")
    normalized["column"] = pd.to_numeric(normalized["column"], errors="coerce")

    invalid_types: list[str] = []
    orphan_rows: list[int] = []

    for index, row in normalized.iterrows():
        parsed_from_id: tuple[str, int, int] | None = None
        decision_id = row["decision_id"]

        if decision_id:
            try:
                parsed_from_id = _parse_decision_id(decision_id)
            except ValueError:
                invalid_types.append(f"Fila {index + 2}: decision_id inválido")
                continue

        row_value = row["row"]
        column_value = row["column"]
        row_has_value = pd.notna(row_value)
        column_has_value = pd.notna(column_value)
        raw_row_value = raw_row_values.iloc[index]
        raw_column_value = raw_column_values.iloc[index]

        row_was_provided = pd.notna(raw_row_value) and str(raw_row_value).strip() != ""
        column_was_provided = pd.notna(raw_column_value) and str(raw_column_value).strip() != ""

        if row_was_provided and not row_has_value:
            invalid_types.append(f"Fila {index + 2}: row debe ser un entero")
            continue
        if column_was_provided and not column_has_value:
            invalid_types.append(f"Fila {index + 2}: column debe ser un entero")
            continue

        if row_has_value and float(row_value) != int(row_value):
            invalid_types.append(f"Fila {index + 2}: row debe ser un entero")
            continue
        if column_has_value and float(column_value) != int(column_value):
            invalid_types.append(f"Fila {index + 2}: column debe ser un entero")
            continue
        if row_has_value and int(row_value) < 1:
            invalid_types.append(f"Fila {index + 2}: row debe ser >= 1")
            continue
        if column_has_value and int(column_value) < 1:
            invalid_types.append(f"Fila {index + 2}: column debe ser >= 1")
            continue

        has_coordinates = bool(row["sheet"]) and row_has_value and column_has_value

        if parsed_from_id and has_coordinates:
            expected_sheet, expected_row, expected_column = parsed_from_id
            if str(row["sheet"]) != expected_sheet or int(row_value) != expected_row or int(column_value) != expected_column:
                orphan_rows.append(index + 2)
                continue

        if not has_coordinates:
            if not parsed_from_id:
                orphan_rows.append(index + 2)
                continue
            normalized.at[index, "sheet"] = parsed_from_id[0]
            normalized.at[index, "row"] = parsed_from_id[1]
            normalized.at[index, "column"] = parsed_from_id[2]

        row_index = int(normalized.at[index, "row"])
        column_index = int(normalized.at[index, "column"])
        coordinate = f"{column_letter(column_index)}{row_index}"
        normalized.at[index, "cell"] = normalized.at[index, "cell"] or coordinate
        normalized.at[index, "column_letter"] = normalized.at[index, "column_letter"] or column_letter(column_index)
        normalized.at[index, "context"] = normalized.at[index, "context"] or (
            f"{normalized.at[index, 'sheet']}!{coordinate} · fila {row_index} · columna {column_letter(column_index)} ({column_index})"
        )

    if invalid_types:
        raise ValueError(f"Tipos inválidos en decisiones: {invalid_types}")

    if orphan_rows:
        raise ValueError(f"Filas huérfanas o inconsistentes en decisiones: {orphan_rows}")

    normalized["row"] = normalized["row"].astype(int)
    normalized["column"] = normalized["column"].astype(int)
    return normalized



def apply_decisions(
    workbook_a: str | Path,
    decisions: pd.DataFrame,
    output_path: str | Path,
    workbook_b: str | Path,
    base: WorkbookSide = "a",
    include_sheets_from_source_only: bool = True,
    compare_mode: CompareMode = "coordinate",
    header_row: int = 1,
    sheet_keys: Mapping[str, Sequence[str]] | None = None,
) -> Path:
    """Aplica decisiones sobre el libro base y genera un archivo combinado.

    Contrato de entrada:
    - `decisions` debe cumplir `validate_decisions_dataframe`.
    - `action` válidas: `use_a`, `use_b`, `manual`.
    - `base="a"` construye el resultado sobre A; `base="b"` sobre B.
    - `compare_mode="row-based"` aplica decisiones por identidad lógica
      (`sheet`, `key`, `header`, `diff_type`) y usa `header_row` / `sheet_keys`
      para localizar registros en ambos libros.

    Salida:
    - `Path` del archivo generado.
    """

    normalized_decisions = validate_decisions_dataframe(decisions)
    compare_mode = decisions.attrs.get("compare_mode", compare_mode)
    header_row = int(decisions.attrs.get("header_row", header_row))
    merged_sheet_keys = {sheet: list(columns) for sheet, columns in (sheet_keys or {}).items()}
    for sheet, columns in decisions.attrs.get("sheet_keys", {}).items():
        merged_sheet_keys.setdefault(sheet, list(columns))
    base_key, source_key = _resolve_direction(base)
    workbook_paths = {"a": workbook_a, "b": workbook_b}

    wb_out = load_workbook(workbook_paths[base_key])
    wb_a = load_workbook(workbook_a)
    wb_b = load_workbook(workbook_b)
    workbooks = {"a": wb_a, "b": wb_b}
    wb_source = workbooks[source_key]

            if action == f"use_{source_key}":
                if sheet_name in wb_source.sheetnames:
                    ws_out.cell(row=row_index, column=column_index).value = wb_source[sheet_name].cell(
                        row=row_index,
                        column=column_index,
                    ).value
                continue

            ws_out.cell(row=row_index, column=column_index).value = row.get("manual_value")

    if include_sheets_from_source_only:
        for sheet in wb_source.sheetnames:
            if sheet in wb_out.sheetnames:
                continue
            src = wb_source[sheet]
            ws_new = wb_out.create_sheet(sheet)
            for row in src.iter_rows():
                for cell in row:
                    ws_new[cell.coordinate] = cell.value

    output = Path(output_path)
    wb_out.save(output)
    return output


def _parse_key_label(key_label: object) -> dict[str, str]:
    if key_label is None:
        return {}
    text = str(key_label).strip()
    if not text or text.startswith("fila="):
        return {}

    parsed: dict[str, str] = {}
    for part in text.split(", "):
        if "=" not in part:
            return {}
        header, value = part.split("=", 1)
        parsed[header.strip()] = value
    return parsed


def _sheet_headers(ws: Worksheet, header_row: int) -> tuple[list[str], dict[str, int]]:
    raw_headers = [ws.cell(row=header_row, column=col).value for col in range(1, ws.max_column + 1)]
    headers: list[str] = []
    header_to_column: dict[str, int] = {}
    for idx, raw_header in enumerate(raw_headers, start=1):
        header = str(raw_header).strip() if raw_header is not None else f"__col_{idx}"
        if not header:
            header = f"__col_{idx}"
        if header in header_to_column:
            header = f"{header}__{idx}"
        headers.append(header)
        header_to_column[header] = idx
    return headers, header_to_column


def _find_row_by_key_label(ws: Worksheet, key_label: object, header_row: int) -> int | None:
    parsed_key = _parse_key_label(key_label)
    if parsed_key:
        headers, _ = _sheet_headers(ws, header_row)
        for row_idx in range(header_row + 1, ws.max_row + 1):
            values = {header: ws.cell(row=row_idx, column=col_idx).value for col_idx, header in enumerate(headers, start=1)}
            if all(_stringify_key_part(values.get(header)) == expected for header, expected in parsed_key.items()):
                return row_idx

    text = str(key_label).strip() if key_label is not None else ""
    if text.startswith("fila="):
        try:
            return int(text.split("=", 1)[1])
        except ValueError:
            return None
    return None


def _ensure_header_column(ws: Worksheet, header: object, header_row: int, fallback_column: int) -> int:
    if header:
        _, header_to_column = _sheet_headers(ws, header_row)
        header_name = str(header)
        if header_name in header_to_column:
            return int(header_to_column[header_name])
        new_column = ws.max_column + 1
        ws.cell(row=header_row, column=new_column).value = header_name
        return new_column
    return max(1, int(fallback_column))


def _copy_row_values(ws_from: Worksheet, row_from: int, ws_to: Worksheet, row_to: int, header_row: int) -> None:
    headers, header_to_column = _sheet_headers(ws_from, header_row)
    for header in headers:
        from_column = header_to_column[header]
        to_column = _ensure_header_column(ws_to, header, header_row, from_column)
        ws_to.cell(row=row_to, column=to_column).value = ws_from.cell(row=row_from, column=from_column).value


def _allocate_row_for_key(ws: Worksheet, key_label: object, header_row: int) -> int:
    existing_row = _find_row_by_key_label(ws, key_label, header_row)
    if existing_row is not None:
        return existing_row

    candidate = max(header_row + 1, ws.max_row + 1)
    parsed_key = _parse_key_label(key_label)
    for header, value in parsed_key.items():
        column_index = _ensure_header_column(ws, header, header_row, ws.max_column + 1)
        ws.cell(row=candidate, column=column_index).value = None if value == "<vacío>" else value
    return candidate


def _apply_row_based_decisions(
    *,
    wb_out: Workbook,
    wb_a: Workbook,
    wb_b: Workbook,
    decisions: pd.DataFrame,
    header_row: int,
    sheet_keys: Mapping[str, Sequence[str]],
) -> None:
    grouped = decisions.groupby(["sheet", "key", "diff_type"], dropna=False, sort=False)

    for (sheet_name, key_label, diff_type), group in grouped:
        sheet_name = str(sheet_name)
        key_label = None if pd.isna(key_label) else key_label
        diff_type = str(diff_type)

        if sheet_name not in wb_out.sheetnames:
            wb_out.create_sheet(sheet_name)
        ws_out = wb_out[sheet_name]

        ws_a = wb_a[sheet_name] if sheet_name in wb_a.sheetnames else None
        ws_b = wb_b[sheet_name] if sheet_name in wb_b.sheetnames else None

        row_out = _find_row_by_key_label(ws_out, key_label, header_row)
        row_a = _find_row_by_key_label(ws_a, key_label, header_row) if ws_a is not None else None
        row_b = _find_row_by_key_label(ws_b, key_label, header_row) if ws_b is not None else None

        if (
            diff_type in {"added", "deleted"}
            and len(set(group["action"].tolist())) == 1
            and group["action"].iloc[0] in {"use_a", "use_b"}
        ):
            action = str(group["action"].iloc[0])
            chosen_sheet = ws_a if action == "use_a" else ws_b
            chosen_row = row_a if action == "use_a" else row_b
            if chosen_sheet is None or chosen_row is None:
                if row_out is not None and diff_type == "deleted":
                    ws_out.delete_rows(row_out, 1)
                continue
            if row_out is None:
                row_out = _allocate_row_for_key(ws_out, key_label, header_row)
            _copy_row_values(chosen_sheet, chosen_row, ws_out, row_out, header_row)
            continue

        if diff_type == "added" and row_out is None:
            row_out = _allocate_row_for_key(ws_out, key_label, header_row)

        for _, decision in group.iterrows():
            action = str(decision["action"])
            header = decision["header"]
            column_index = _ensure_header_column(ws_out, header, header_row, int(decision["column"]))

            if row_out is None:
                fallback_label = f"fila={int(decision['row'])}"
                row_out = _allocate_row_for_key(ws_out, fallback_label, header_row)

            if action == "manual":
                ws_out.cell(row=row_out, column=column_index).value = decision.get("manual_value")
                continue

            if action == "use_a":
                if ws_a is None or row_a is None:
                    ws_out.cell(row=row_out, column=column_index).value = None
                    continue
                source_column = _ensure_header_column(ws_a, header, header_row, int(decision["column"]))
                ws_out.cell(row=row_out, column=column_index).value = ws_a.cell(row=row_a, column=source_column).value
                continue

            if ws_b is None or row_b is None:
                ws_out.cell(row=row_out, column=column_index).value = None
                continue
            source_column = _ensure_header_column(ws_b, header, header_row, int(decision["column"]))
            ws_out.cell(row=row_out, column=column_index).value = ws_b.cell(row=row_b, column=source_column).value


__all__ = [
    "CellDiff",
    "CompareOptions",
    "ComparisonRequest",
    "ComparatorService",
    "DECISION_COLUMNS",
    "DECISION_FORMAT_VERSION",
    "DECISION_METADATA_SHEET_NAME",
    "DECISION_SHEET_NAME",
    "DEFAULT_ACTION",
    "DecisionTemplateRequest",
    "DecisionsLoadRequest",
    "MergeRequest",
    "SERVICE",
    "VALID_ACTIONS",
    "VALID_COMPARE_MODES",
    "WorkbookDiff",
    "WorkbookSide",
    "apply_decisions",
    "build_decision_id",
    "column_letter",
    "compare_workbooks",
    "decisions_from_excel",
    "diffs_to_dataframe",
    "export_decision_template",
    "source_action_for_base",
    "validate_decisions_dataframe",
]
