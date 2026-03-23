from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from comparator import (
    CompareOptions,
    ComparisonRequest,
    DECISION_COLUMNS,
    DECISION_FORMAT_VERSION,
    DECISION_METADATA_SHEET_NAME,
    MergeRequest,
    SERVICE,
    apply_decisions,
    compare_workbooks,
    decisions_from_excel,
    diffs_to_dataframe,
    export_decision_template,
    source_action_for_base,
    validate_decisions_dataframe,
)
from cli_adapter import parse_sheet_keys_args
from interface_adapter import build_compare_options
from streamlit_adapter import build_review_table, parse_sheet_keys_block


def _create_wb(path: Path, sheets: dict[str, dict[str, object]]):
    wb = Workbook()
    wb.remove(wb.active)
    for sheet, cells in sheets.items():
        ws = wb.create_sheet(sheet)
        for coord, value in cells.items():
            ws[coord] = value
    wb.save(path)



def test_compare_workbooks_detects_sheet_and_cell_differences(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"

    _create_wb(
        a,
        {
            "Resumen": {"A1": "ID", "A2": 1, "B2": "Ana"},
            "SoloA": {"A1": "x"},
        },
    )
    _create_wb(
        b,
        {
            "Resumen": {"A1": "ID", "A2": 1, "B2": "Anita"},
            "SoloB": {"A1": "y"},
        },
    )

    diff = compare_workbooks(a, b)

    assert diff.only_in_a == ["SoloA"]
    assert diff.only_in_b == ["SoloB"]
    assert diff.common_sheets == ["Resumen"]
    assert len(diff.differences["Resumen"]) == 1
    assert diff.differences["Resumen"][0].coordinate == "B2"



def test_compare_options_ignore_case_and_spaces(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"

    _create_wb(a, {"Datos": {"A1": "  Ana  "}})
    _create_wb(b, {"Datos": {"A1": "ana"}})

    default_diff = compare_workbooks(a, b)
    assert len(default_diff.all_differences()) == 1

    relaxed_diff = compare_workbooks(
        a,
        b,
        options=CompareOptions(strip_strings=True, case_sensitive=False),
    )
    assert relaxed_diff.all_differences() == []



def test_compare_workbooks_row_based_avoids_cascade_and_detects_row_states(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"

    _create_wb(
        a,
        {
            "Datos": {
                "A1": "ID",
                "B1": "Nombre",
                "C1": "Estado",
                "A2": 1,
                "B2": "Ana",
                "C2": "OK",
                "A3": 2,
                "B3": "Luis",
                "C3": "OK",
            }
        },
    )
    _create_wb(
        b,
        {
            "Datos": {
                "A1": "ID",
                "B1": "Nombre",
                "C1": "Estado",
                "A2": 1,
                "B2": "Ana",
                "C2": "Actualizado",
                "A3": 3,
                "B3": "Marta",
                "C3": "Nueva",
                "A4": 2,
                "B4": "Luis",
                "C4": "OK",
            }
        },
    )

    coordinate_diff = compare_workbooks(a, b)
    assert len(coordinate_diff.all_differences()) == 7

    row_diff = compare_workbooks(
        a,
        b,
        options=CompareOptions(compare_mode="row-based", sheet_keys={"Datos": ["ID"]}),
    )

    row_diffs = row_diff.differences["Datos"]
    assert len(row_diffs) == 4
    assert {diff.diff_type for diff in row_diffs} == {"modified", "added"}
    modified = [diff for diff in row_diffs if diff.diff_type == "modified"]
    added = [diff for diff in row_diffs if diff.diff_type == "added"]
    assert len(modified) == 1
    assert modified[0].header == "Estado"
    assert modified[0].key == "ID=1"
    assert {diff.header for diff in added} == {"ID", "Nombre", "Estado"}
    assert {diff.key for diff in added} == {"ID=3"}



def test_diffs_to_dataframe_exposes_stable_contract_columns_and_decision_ids(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"

    _create_wb(a, {"Datos": {"A1": "x"}})
    _create_wb(b, {"Datos": {"A1": "y"}})

    diff = compare_workbooks(a, b)
    df = diffs_to_dataframe(diff.all_differences())

    assert list(df.columns) == DECISION_COLUMNS
    assert df.iloc[0]["decision_id"] == "Datos|A1|modified"
    assert df.iloc[0]["diff_type"] == "modified"
    assert df.iloc[0]["decision_id"].startswith("Datos|1|1|")



def test_apply_decisions_merges_changes(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"

    _create_wb(a, {"Datos": {"A1": "hola", "A2": "base"}})
    _create_wb(b, {"Datos": {"A1": "hola", "A2": "nuevo"}, "Nueva": {"A1": 99}})

    diff = compare_workbooks(a, b)
    df = diffs_to_dataframe(diff.all_differences(), default_action="use_b")

    output = tmp_path / "out.xlsx"
    apply_decisions(a, df, output, b, base="a")

    wb = load_workbook(output)
    assert wb["Datos"]["A2"].value == "nuevo"
    assert wb["Nueva"]["A1"].value == 99



def test_apply_decisions_merges_changes_from_a_onto_b(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"

    _create_wb(a, {"Datos": {"A1": "hola", "A2": "nuevo"}, "NuevaA": {"A1": 42}})
    _create_wb(b, {"Datos": {"A1": "hola", "A2": "base"}})

    diff = compare_workbooks(a, b)
    df = diffs_to_dataframe(diff.all_differences(), default_action="use_a")

    output = tmp_path / "out_b.xlsx"
    apply_decisions(a, df, output, b, base="b")

    wb = load_workbook(output)
    assert wb["Datos"]["A2"].value == "nuevo"
    assert wb["NuevaA"]["A1"].value == 42



def test_export_and_read_decisions_template(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"

    _create_wb(a, {"Datos": {"A1": "x"}})
    _create_wb(b, {"Datos": {"A1": "y"}})

    diff = compare_workbooks(a, b)
    template = tmp_path / "decisiones.xlsx"
    export_decision_template(diff, template)

    loaded = decisions_from_excel(template)
    assert loaded.shape[0] == 1
    assert loaded.iloc[0]["action"] == "use_b"
    assert loaded.iloc[0]["sheet"] == "Datos"
    assert loaded.iloc[0]["decision_id"] == "Datos|A1|modified"
    assert loaded.iloc[0]["diff_type"] == "modified"
    assert loaded.iloc[0]["decision_id"].startswith("Datos|1|1|")



def test_export_template_protects_structure_adds_table_and_metadata(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    template = tmp_path / "decisiones.xlsx"

    _create_wb(a, {"Datos": {"A1": "base"}})
    _create_wb(b, {"Datos": {"A1": "cambio"}})

    diff = compare_workbooks(a, b)
    export_decision_template(diff, template, default_action="use_b")

    wb = load_workbook(template)
    ws = wb["Decisiones"]
    metadata = wb[DECISION_METADATA_SHEET_NAME]

    decision_id_col = DECISION_COLUMNS.index("decision_id") + 1
    action_col = DECISION_COLUMNS.index("action") + 1
    manual_col = DECISION_COLUMNS.index("manual_value") + 1
    reviewed_col = DECISION_COLUMNS.index("reviewed") + 1

    assert ws.protection.sheet is True
    assert ws.cell(row=2, column=decision_id_col).protection.locked is True
    assert ws.cell(row=2, column=action_col).protection.locked is False
    assert ws.cell(row=2, column=manual_col).protection.locked is False
    assert ws.cell(row=2, column=reviewed_col).protection.locked is False
    assert len(ws.tables) == 1
    assert "DecisionTable" in ws.tables
    assert metadata.sheet_state == "hidden"

    metadata_values = {row[0]: row[1] for row in metadata.iter_rows(min_row=2, values_only=True)}
    assert metadata_values["format_version"] == DECISION_FORMAT_VERSION
    assert metadata_values["compare_mode"] == "coordinate"
    assert metadata_values["base_workbook"] == "a"
    assert metadata_values["workbook_a"] == "a.xlsx"
    assert metadata_values["workbook_b"] == "b.xlsx"
    assert metadata_values["source_signature"]
    assert metadata_values["generated_at"]



def test_decisions_validation_supports_decision_id_without_coordinates(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    output = tmp_path / "out.xlsx"

    _create_wb(a, {"Datos": {"A1": "base"}})
    _create_wb(b, {"Datos": {"A1": "source"}})

    diff = compare_workbooks(a, b)
    df = diffs_to_dataframe(diff.all_differences(), default_action="use_b")
    df.loc[0, ["sheet", "row", "column", "cell", "column_letter", "context"]] = [None, None, None, None, None, None]

    apply_decisions(a, df, output, b, base="a")

    wb = load_workbook(output)
    assert wb["Datos"]["A1"].value == "source"



def test_decisions_from_excel_detects_deleted_columns(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    template = tmp_path / "decisiones.xlsx"

    _create_wb(a, {"Datos": {"A1": "x"}})
    _create_wb(b, {"Datos": {"A1": "y"}})

    diff = compare_workbooks(a, b)
    export_decision_template(diff, template)

    wb = load_workbook(template)
    ws = wb["Decisiones"]
    ws.delete_cols(1)
    wb.save(template)

    with pytest.raises(ValueError, match="columnas requeridas"):
        decisions_from_excel(template)



def test_validate_decisions_dataframe_rejects_invalid_actions():
    invalid = diffs_to_dataframe([])
    invalid.loc[0] = [None] * len(DECISION_COLUMNS)
    invalid.loc[0, "decision_id"] = "Datos|1|1|abcdef123456"
    invalid.loc[0, "sheet"] = "Datos"
    invalid.loc[0, "row"] = 1
    invalid.loc[0, "column"] = 1
    invalid.loc[0, "action"] = "romper"

    with pytest.raises(ValueError, match="Acciones no válidas"):
        validate_decisions_dataframe(invalid)



def test_validate_decisions_dataframe_rejects_invalid_types_duplicate_ids_and_orphans(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"

    _create_wb(a, {"Datos": {"A1": "x", "A2": "base"}})
    _create_wb(b, {"Datos": {"A1": "y", "A2": "source"}})

    diff = compare_workbooks(a, b)
    df = diffs_to_dataframe(diff.all_differences())

    duplicate = pd.concat([df.iloc[[0]], df.iloc[[0]]], ignore_index=True)
    with pytest.raises(ValueError, match="decision_id duplicados"):
        validate_decisions_dataframe(duplicate)

    invalid_type = df.copy()
    invalid_type["row"] = invalid_type["row"].astype(object)
    invalid_type.loc[0, "row"] = "fila-uno"
    with pytest.raises(ValueError, match="Tipos inválidos"):
        validate_decisions_dataframe(invalid_type)

    orphan = df.copy()
    orphan.loc[0, "sheet"] = "OtraHoja"
    with pytest.raises(ValueError, match="Filas huérfanas"):
        validate_decisions_dataframe(orphan)



def test_decisions_from_excel_preserves_row_based_metadata(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"

    _create_wb(a, {"Datos": {"A1": "ID", "B1": "Nombre", "A2": 1, "B2": "Ana"}})
    _create_wb(b, {"Datos": {"A1": "ID", "B1": "Nombre", "A2": 2, "B2": "Luis"}})

    diff = compare_workbooks(a, b, options=CompareOptions(compare_mode="row-based", sheet_keys={"Datos": ["ID"]}))
    template = tmp_path / "metadata.xlsx"
    export_decision_template(diff, template)

    loaded = decisions_from_excel(template)
    assert loaded.attrs["compare_mode"] == "row-based"
    assert loaded.attrs["header_row"] == 1
    assert loaded.attrs["sheet_keys"] == {"Datos": ["ID"]}


def test_excel_round_trip_supports_manual_override_and_merge(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    template = tmp_path / "decisiones.xlsx"
    output = tmp_path / "merge.xlsx"

    _create_wb(a, {"Datos": {"A1": "ID", "B1": "Estado", "A2": 1, "B2": "base"}})
    _create_wb(b, {"Datos": {"A1": "ID", "B1": "Estado", "A2": 1, "B2": "nuevo"}})

    diff = compare_workbooks(a, b)
    export_decision_template(diff, template)

    wb_template = load_workbook(template)
    ws = wb_template["Decisiones"]
    action_column = DECISION_COLUMNS.index("action") + 1
    manual_column = DECISION_COLUMNS.index("manual_value") + 1
    ws.cell(row=2, column=action_column).value = "manual"
    ws.cell(row=2, column=manual_column).value = "forzado"
    wb_template.save(template)

    decisions = decisions_from_excel(template)
    assert decisions.iloc[0]["action"] == "manual"
    assert decisions.iloc[0]["manual_value"] == "forzado"

    apply_decisions(a, decisions, output, b, base="a")

    wb_out = load_workbook(output)
    assert wb_out["Datos"]["B2"].value == "forzado"


def test_decision_template_exports_excel_controls_and_protection(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    template = tmp_path / "decisiones.xlsx"

    _create_wb(a, {"Datos": {"A1": "x"}})
    _create_wb(b, {"Datos": {"A1": "y"}})

    export_decision_template(compare_workbooks(a, b), template)

    ws = load_workbook(template)["Decisiones"]
    validations = list(ws.data_validations.dataValidation)

    assert ws.freeze_panes == "A2"
    assert ws.protection.sheet is True
    assert len(validations) == 1
    assert validations[0].formula1 == '"use_a,use_b,manual"'
    assert validations[0].sqref == "M2:M2"
    assert ws["M2"].protection.locked is False
    assert ws["N2"].protection.locked is False
    assert ws["O2"].protection.locked is False
    assert ws["A2"].protection.locked is True


def test_decisions_from_excel_rejects_missing_required_column(tmp_path: Path):
    template = tmp_path / "decisiones.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Decisiones"
    ws.append([column for column in DECISION_COLUMNS if column != "decision_id"])
    ws.append(["Datos", "A1", 1, 1, "A", "ctx", "h", "k", "modified", "x", "y", "use_b", None, False])
    wb.save(template)

    with pytest.raises(ValueError, match="columnas requeridas"):
        decisions_from_excel(template)


def test_decisions_from_excel_rejects_duplicate_decision_id(tmp_path: Path):
    template = tmp_path / "decisiones.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Decisiones"
    ws.append(DECISION_COLUMNS)
    row = [
        "Datos",
        "A1",
        "repetido",
        1,
        1,
        "A",
        "ctx",
        "h",
        "k",
        "modified",
        "x",
        "y",
        "use_b",
        None,
        False,
    ]
    ws.append(row)
    ws.append(row)
    wb.save(template)

    with pytest.raises(ValueError, match="decision_id duplicados"):
        decisions_from_excel(template)


@pytest.mark.parametrize(
    ("column_name", "value", "message"),
    [
        ("action", "romper", "Acciones no válidas"),
        ("row", "fila-dos", "columna 'row'"),
        ("sheet", "   ", "hoja válida"),
        ("decision_id", None, "decision_id"),
    ],
)
def test_validate_decisions_dataframe_rejects_controlled_corruption(column_name: str, value: object, message: str):
    invalid = diffs_to_dataframe([])
    invalid.loc[0] = [None] * len(DECISION_COLUMNS)
    invalid.loc[0, "sheet"] = "Datos"
    invalid.loc[0, "cell"] = "A1"
    invalid.loc[0, "decision_id"] = "Datos|A1|modified"
    invalid.loc[0, "row"] = 1
    invalid.loc[0, "column"] = 1
    invalid.loc[0, "action"] = "use_b"
    invalid.loc[0, column_name] = value

    with pytest.raises(ValueError, match=message):
        validate_decisions_dataframe(invalid)


def test_service_api_supports_compare_and_merge_requests(tmp_path: Path):
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    out = tmp_path / "merged.xlsx"

    _create_wb(a, {"Datos": {"A1": "base"}})
    _create_wb(b, {"Datos": {"A1": "source"}})

    diff = SERVICE.compare(ComparisonRequest(path_a=a, path_b=b))
    decisions = diff.to_dataframe(default_action=source_action_for_base("a"))
    result = SERVICE.apply_decisions(
        MergeRequest(
            workbook_a=a,
            workbook_b=b,
            decisions=decisions,
            output_path=out,
            base="a",
        )
    )

    wb = load_workbook(result)
    assert wb["Datos"]["A1"].value == "source"


def test_interface_adapter_parsers_and_option_builder():
    assert parse_sheet_keys_block("Clientes:ID\nPedidos:Empresa,Numero") == {
        "Clientes": ["ID"],
        "Pedidos": ["Empresa", "Numero"],
    }
    assert parse_sheet_keys_args(["Clientes=ID", "Pedidos=Empresa,Numero"]) == {
        "Clientes": ["ID"],
        "Pedidos": ["Empresa", "Numero"],
    }

    options = build_compare_options(
        compare_mode="row-based",
        ignore_case=True,
        keep_spaces=False,
        empty_string_is_value=False,
        header_row=2,
        sheet_keys={"Clientes": ["ID"]},
    )
    assert options.compare_mode == "row-based"
    assert options.case_sensitive is False
    assert options.header_row == 2


def test_excel_addin_adapter_supports_compare_sheet_roundtrip_and_merge(tmp_path: Path):
    adapter = ExcelAddinAdapter()
    base = tmp_path / "base.xlsx"
    source = tmp_path / "source.xlsx"
    host = tmp_path / "host.xlsx"
    output = tmp_path / "merged_from_adapter.xlsx"

    _create_wb(base, {"Datos": {"A1": "hola", "A2": "base"}})
    _create_wb(source, {"Datos": {"A1": "hola", "A2": "source"}, "Nueva": {"A1": 7}})
    _create_wb(host, {"Inicio": {"A1": "placeholder"}})

    selection = adapter.select_workbooks(
        base_workbook_path=base,
        source_workbook_path=source,
        base_side="a",
    )
    comparison = adapter.compare(ExcelCompareContract(selection=selection))

    assert comparison.route == adapter.integration_route
    assert comparison.only_in_base == []
    assert comparison.only_in_source == ["Nueva"]
    assert comparison.decision_rows[0].action == "use_b"

    sheet_contract = ExcelDecisionSheetContract.create(host, sheet_name="DecisionTable")
    load_result = adapter.load_decision_table_into_workbook(comparison, sheet_contract)
    assert load_result.rows_written == 1

    loaded = adapter.read_decisions_from_workbook(sheet_contract)
    assert loaded.rows_loaded == 1
    assert loaded.decisions[0].sheet == "Datos"

    merge_result = adapter.execute_merge(
        ExcelMergeContract.create(
            selection=selection,
            decisions_workbook_path=host,
            decisions_sheet_name="DecisionTable",
            output_path=output,
        )
    )
    merged = load_workbook(merge_result.output_path)
    assert merged["Datos"]["A2"].value == "source"
    assert merged["Nueva"]["A1"].value == 7


def test_excel_addin_adapter_payload_contract_is_serializable(tmp_path: Path):
    adapter = ExcelAddinAdapter()
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    decisions_host = tmp_path / "decisions_host.xlsx"
    out = tmp_path / "out.xlsx"

    _create_wb(a, {"Datos": {"A1": "uno"}})
    _create_wb(b, {"Datos": {"A1": "dos"}})
    _create_wb(decisions_host, {"Inicio": {"A1": "x"}})

    comparison_payload = adapter.compare_payload(
        {
            "selection": {
                "base_workbook_path": str(a),
                "source_workbook_path": str(b),
                "base_side": "a",
            },
            "compare_mode": "coordinate",
        }
    )
    assert comparison_payload["total_differences"] == 1
    assert comparison_payload["decision_rows"][0]["action"] == "use_b"

    load_payload = adapter.load_decision_table_payload(
        comparison_payload,
        {"workbook_path": str(decisions_host), "sheet_name": "DecisionTable"},
    )
    assert load_payload["rows_written"] == 1

    decisions_payload = adapter.read_decisions_payload(
        {"workbook_path": str(decisions_host), "sheet_name": "DecisionTable"}
    )
    assert decisions_payload["rows_loaded"] == 1

    merge_payload = adapter.execute_merge_payload(
        {
            "selection": {
                "base_workbook_path": str(a),
                "source_workbook_path": str(b),
                "base_side": "a",
            },
            "decisions_workbook_path": str(decisions_host),
            "decisions_sheet_name": "DecisionTable",
            "output_path": str(out),
        }
    )
    wb = load_workbook(merge_payload["output_path"])
    assert wb["Datos"]["A1"].value == "dos"
